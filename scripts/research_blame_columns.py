#!/usr/bin/env python3
"""Fill columns R/S (blame) in delivery research export; send xlsx to owner."""

from __future__ import annotations

import json
import os
import sys
import time
import urllib.parse
import urllib.request
from collections import defaultdict
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("pip install openpyxl", file=sys.stderr)
    raise SystemExit(1)

INBOX = Path("/root/research/inbox")
OUT_DIR = Path("/root/research/out")
ENV_FILE = Path("/root/.video_bot.env")
BATCH_WINDOW_MIN = int(os.environ.get("RESEARCH_BATCH_WINDOW_MIN", "60"))
CTE_OK_MAX = float(os.environ.get("RESEARCH_CTE_OK_MAX", "35"))
DISPATCH_SEARCH_MIN = float(os.environ.get("RESEARCH_DISPATCH_SEARCH_MIN", "15"))


def load_env() -> dict[str, str]:
    env: dict[str, str] = {}
    for key, value in os.environ.items():
        if key.startswith("TG_"):
            env[key] = value
    if ENV_FILE.exists():
        for raw in ENV_FILE.read_text(encoding="utf-8").splitlines():
            line = raw.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            key, _, value = line.partition("=")
            env[key.strip()] = value.strip().strip('"').strip("'")
    return env


def tg_send_text(text: str, env: dict[str, str]) -> None:
    token = env.get("TG_BOT_TOKEN", "")
    chat = env.get("TG_CHAT_ID", "")
    if not token or not chat:
        return
    data = urllib.parse.urlencode({"chat_id": chat, "text": text[:3900]}).encode()
    req = urllib.request.Request(
        f"https://api.telegram.org/bot{token}/sendMessage",
        data=data,
        method="POST",
    )
    with urllib.request.urlopen(req, timeout=60) as resp:
        json.loads(resp.read().decode())


def tg_send_document(path: Path, caption: str, env: dict[str, str]) -> None:
    token = env.get("TG_BOT_TOKEN", "")
    chat = env.get("TG_CHAT_ID", "")
    if not token or not chat:
        raise RuntimeError("TG_BOT_TOKEN/TG_CHAT_ID missing")
    boundary = f"----boundary{int(time.time())}"
    body = bytearray()
    for name, val in (("chat_id", chat), ("caption", caption[:1024])):
        body.extend(f"--{boundary}\r\n".encode())
        body.extend(f'Content-Disposition: form-data; name="{name}"\r\n\r\n'.encode())
        body.extend(f"{val}\r\n".encode())
    body.extend(f"--{boundary}\r\n".encode())
    body.extend(
        f'Content-Disposition: form-data; name="document"; filename="{path.name}"\r\n'.encode()
    )
    body.extend(b"Content-Type: application/vnd.openxmlformats-officedocument.spreadsheetml.sheet\r\n\r\n")
    body.extend(path.read_bytes())
    body.extend(f"\r\n--{boundary}--\r\n".encode())
    req = urllib.request.Request(
        f"https://api.telegram.org/bot{token}/sendDocument",
        data=bytes(body),
        method="POST",
    )
    req.add_header("Content-Type", f"multipart/form-data; boundary={boundary}")
    with urllib.request.urlopen(req, timeout=600) as resp:
        result = json.loads(resp.read().decode())
    if not result.get("ok"):
        raise RuntimeError(f"sendDocument failed: {result}")


def parse_dt(value) -> datetime | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, str):
        text = value.strip()
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%d.%m.%Y %H:%M:%S"):
            try:
                return datetime.strptime(text[:19], fmt)
            except ValueError:
                continue
    return None


def parse_float(value) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(str(value).replace(",", "."))
    except ValueError:
        return None


def parse_int(value) -> int:
    try:
        return int(float(str(value).replace(",", ".")))
    except (TypeError, ValueError):
        return 0


def delta_minutes(start: datetime | None, end: datetime | None) -> float | None:
    if start is None or end is None:
        return None
    sec = (end - start).total_seconds()
    if sec < 0:
        return None
    return sec / 60.0


def round_min(value: float | None) -> int:
    if value is None:
        return 0
    return int(round(value))


def compute_batches(rows: list[dict]) -> list[int]:
    """Batch size minus self: same courier (AA), store (B), same calendar day, |N-N| <= window."""
    groups: dict[tuple, list[tuple[int, datetime]]] = defaultdict(list)
    for idx, row in enumerate(rows):
        picked = row.get("picked")
        store = row.get("store")
        courier = row.get("courier")
        if picked is None or not store or not courier:
            continue
        key = (str(store), str(courier), picked.date())
        groups[key].append((idx, picked))

    batch_extra = [0] * len(rows)
    for items in groups.values():
        items.sort(key=lambda x: x[1])
        for i, (idx_a, t_a) in enumerate(items):
            count = 0
            for j, (_, t_b) in enumerate(items):
                if i == j:
                    continue
                if abs((t_b - t_a).total_seconds()) / 60.0 <= BATCH_WINDOW_MIN:
                    count += 1
            batch_extra[idx_a] = count
    return batch_extra


def classify_row(row: dict, batch_extra: int) -> tuple[str, str]:
    cte = row.get("cte")
    if cte is not None and cte < CTE_OK_MAX:
        cte_txt = f"{cte:.1f}".rstrip("0").rstrip(".")
        return "Нет проблем", f"CTE {cte_txt} мин — нет проблем"

    picked = row.get("picked")
    delivery_start = row.get("delivery_start")
    assembled = row.get("assembled")
    courier_arrival = row.get("courier_arrival")
    search_start = row.get("search_start")
    courier_assigned = row.get("courier_assigned")
    u_attempts = row.get("u_attempts", 0)
    v_attempts = row.get("v_attempts", 0)

    if picked and not delivery_start:
        return (
            "Курьер",
            "Стоит picked_date, но нет даты начала доставки — заказ не уехал",
        )

    if assembled and courier_arrival and assembled > courier_arrival:
        wait = delta_minutes(courier_arrival, assembled)
        return (
            "Магазин",
            f"Сборку закончили после приезда курьера — он ждал {round_min(wait)} мин",
        )

    if batch_extra >= 1:
        pack_size = batch_extra + 1
        courier = row.get("courier") or "?"
        store = row.get("store") or "?"
        return (
            "Пачка",
            f"Пачка {pack_size} заказов: курьер {courier}, магазин {store}, "
            f"окно {BATCH_WINDOW_MIN} мин от picked",
        )

    picked_to_start = delta_minutes(picked, delivery_start)
    if picked and delivery_start and picked_to_start is not None:
        return (
            "Курьер",
            f"После отметки «выдал» до старта доставки прошло {round_min(picked_to_start)} мин",
        )

    search_min = delta_minutes(search_start, courier_assigned)
    if search_min is not None and search_min > DISPATCH_SEARCH_MIN:
        return (
            "Диспетчеризация",
            f"Поиск курьера {round_min(search_min)} мин "
            f"(попытки 1PL={u_attempts}, 3PL={v_attempts})",
        )

    if not picked and not assembled:
        return "Нет данных", "Недостаточно дат для классификации"

    return "Нет данных", "Недостаточно дат для классификации"


def read_rows(ws) -> list[dict]:
    rows: list[dict] = []
    for excel_row in ws.iter_rows(min_row=2, values_only=True):
        if not excel_row or excel_row[0] in (None, ""):
            continue
        rows.append(
            {
                "cte": parse_float(excel_row[4] if len(excel_row) > 4 else None),
                "store": excel_row[1] if len(excel_row) > 1 else None,
                "assembled": parse_dt(excel_row[8] if len(excel_row) > 8 else None),
                "search_start": parse_dt(excel_row[9] if len(excel_row) > 9 else None),
                "courier_assigned": parse_dt(excel_row[10] if len(excel_row) > 10 else None),
                "courier_arrival": parse_dt(excel_row[12] if len(excel_row) > 12 else None),
                "picked": parse_dt(excel_row[13] if len(excel_row) > 13 else None),
                "delivery_start": parse_dt(excel_row[14] if len(excel_row) > 14 else None),
                "u_attempts": parse_int(excel_row[20] if len(excel_row) > 20 else 0),
                "v_attempts": parse_int(excel_row[21] if len(excel_row) > 21 else 0),
                "courier": excel_row[26] if len(excel_row) > 26 else None,
            }
        )
    return rows


def process_file(src: Path, dst: Path) -> dict:
    print(f"loading {src}", flush=True)
    wb = load_workbook(src)
    ws = wb.active

    parsed = read_rows(ws)
    print(f"rows {len(parsed)}", flush=True)
    batches = compute_batches(parsed)

    stats: dict[str, int] = defaultdict(int)
    for i, row in enumerate(parsed):
        blame, reason = classify_row(row, batches[i])
        stats[blame] += 1
        excel_row_idx = i + 2
        ws.cell(row=excel_row_idx, column=18, value=blame)  # R
        ws.cell(row=excel_row_idx, column=19, value=reason)  # S
        ws.cell(row=excel_row_idx, column=28, value=batches[i])  # AB batch extra
        if i and i % 10000 == 0:
            print(f"  … {i} rows", flush=True)

    ws.cell(row=1, column=28, value="Пачка_доп")
    ws.cell(row=1, column=29, value=f"Окно_мин={BATCH_WINDOW_MIN}")

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    print(f"saving {dst}", flush=True)
    wb.save(dst)
    wb.close()

    ok_cte = sum(1 for r in parsed if r.get("cte") is not None and r["cte"] < CTE_OK_MAX)
    return {
        "rows": len(parsed),
        "cte_ok_lt35": ok_cte,
        "blame": dict(stats),
        "batch_window_min": BATCH_WINDOW_MIN,
        "output": str(dst),
    }


def latest_inbox_xlsx() -> Path | None:
    files = sorted(INBOX.glob("*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
    return files[0] if files else None


def main() -> int:
    env = load_env()
    src = Path(sys.argv[1]) if len(sys.argv) > 1 else latest_inbox_xlsx()
    if not src or not src.exists():
        tg_send_text("Нет .xlsx для заполнения R/S.", env)
        return 1

    stamp = time.strftime("%Y%m%d_%H%M%S")
    dst = OUT_DIR / f"{stamp}_{src.stem}_blame.xlsx"

    try:
        summary = process_file(src, dst)
    except Exception as exc:
        tg_send_text(f"Ошибка заполнения R/S: {exc}", env)
        raise

    blame_lines = "\n".join(
        f"• {k}: {v:,} ({100.0 * v / summary['rows']:.1f}%)".replace(",", " ")
        for k, v in sorted(summary["blame"].items(), key=lambda x: -x[1])
    )
    caption = (
        f"📎 Заполнены R/S (CTE<{CTE_OK_MAX:g} → «Нет проблем»)\n"
        f"Строк: {summary['rows']:,}\n"
        f"CTE OK: {summary['cte_ok_lt35']:,}\n\n"
        f"{blame_lines}"
    ).replace(",", " ")

    tg_send_text(
        f"✅ Готов файл с виновными\n"
        f"Исходник: {src.name}\n"
        f"Строк: {summary['rows']:,}\n\n"
        f"Распределение R:\n{blame_lines}\n\n"
        f"Отправляю файл…",
        env,
    )
    tg_send_document(dst, caption, env)
    print("sent", dst)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
