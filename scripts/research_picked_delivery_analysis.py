#!/usr/bin/env python3
"""
Analyze courier delay: Picked_date → delivery start.

Reads latest .csv or .xlsx from /root/research/inbox/, posts Telegram report.
"""

from __future__ import annotations

import csv
import json
import os
import re
import sys
import time
import urllib.parse
import urllib.request
from collections import defaultdict
from datetime import datetime
from pathlib import Path

INBOX = Path("/root/research/inbox")
OUT = Path("/root/data/mlbb/research_reports")
ENV_FILE = Path("/root/.video_bot.env")

PICKED_ALIASES = (
    "picked_date",
    "picked date",
    "picked",
    "дата сборки",
    "собран",
    "picked_at",
)
DELIVERY_START_ALIASES = (
    "дата начала доставки",
    "дата начала доставки ",
    "delivery_start",
    "delivery start",
    "start_delivery",
    "начало доставки",
    "старт доставки",
)
STORE_ALIASES = (
    "store",
    "store_name",
    "shop",
    "магазин",
    "darkstore",
    "точка",
    "склад",
    "hub",
    "store_id",
)
COURIER_ALIASES = (
    "courier",
    "courier_id",
    "courier_name",
    "курьер",
    "id курьера",
    "courierid",
)
ORDER_ALIASES = ("order_id", "order", "заказ", "id заказа", "orderid")


def load_env() -> dict[str, str]:
    env: dict[str, str] = {}
    for key, value in os.environ.items():
        if key.startswith("TG_"):
            env[key] = value
    if not ENV_FILE.exists():
        return env
    for raw in ENV_FILE.read_text().splitlines():
        line = raw.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, _, value = line.partition("=")
        key, value = key.strip(), value.strip().strip('"').strip("'")
        if key:
            env[key] = value
    return env


def tg_send(text: str, env: dict[str, str] | None = None) -> None:
    env = env or load_env()
    token = env.get("TG_BOT_TOKEN", "")
    chat = env.get("TG_CHAT_ID", "")
    if not token or not chat:
        print(text)
        return
    for chunk_start in range(0, len(text), 3900):
        chunk = text[chunk_start : chunk_start + 3900]
        data = urllib.parse.urlencode({"chat_id": chat, "text": chunk}).encode()
        req = urllib.request.Request(
            f"https://api.telegram.org/bot{token}/sendMessage",
            data=data,
            method="POST",
        )
        with urllib.request.urlopen(req, timeout=30):
            pass


def norm_col(name: str) -> str:
    return re.sub(r"\s+", " ", (name or "").strip().lower())


def find_column(headers: list[str], aliases: tuple[str, ...]) -> str | None:
    normed = {norm_col(h): h for h in headers}
    for alias in aliases:
        if alias in normed:
            return normed[alias]
    for alias in aliases:
        for nk, orig in normed.items():
            if alias in nk or nk in alias:
                return orig
    return None


def parse_dt(value: str | None) -> datetime | None:
    if value is None:
        return None
    s = str(value).strip()
    if not s or s.lower() in ("nan", "none", "null", "nat", ""):
        return None
    s = s.replace("T", " ").replace("Z", "")[:26]
    for fmt in (
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M:%S.%f",
        "%Y-%m-%d %H:%M",
        "%d.%m.%Y %H:%M:%S",
        "%d.%m.%Y %H:%M",
        "%Y/%m/%d %H:%M:%S",
        "%Y-%m-%d",
    ):
        try:
            return datetime.strptime(s[: len(fmt) + 3].strip(), fmt)
        except ValueError:
            continue
    return None


def delay_minutes(picked: datetime | None, delivery: datetime | None) -> float | None:
    if picked is None or delivery is None:
        return None
    sec = (delivery - picked).total_seconds()
    if sec < 0:
        return None
    return sec / 60.0


def pct(n: int, total: int) -> float:
    return 100.0 * n / total if total else 0.0


def read_csv_rows(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    for enc in ("utf-8-sig", "utf-8", "cp1251", "latin-1"):
        try:
            with path.open(encoding=enc, newline="") as handle:
                reader = csv.DictReader(handle)
                headers = list(reader.fieldnames or [])
                rows = list(reader)
            return headers, rows
        except (UnicodeDecodeError, csv.Error):
            continue
    raise ValueError(f"cannot decode csv: {path}")


def read_xlsx_rows(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    it = ws.iter_rows(values_only=True)
    header_row = next(it, None)
    if not header_row:
        wb.close()
        return [], []
    headers = [str(c).strip() if c is not None else "" for c in header_row]
    rows: list[dict[str, str]] = []
    for row in it:
        if not row:
            continue
        d = {headers[i]: ("" if row[i] is None else str(row[i])) for i in range(len(headers))}
        if any(str(v).strip() for v in d.values()):
            rows.append(d)
    wb.close()
    return headers, rows


def top_offenders(
    bucket: dict[str, list[float]],
    *,
    min_n: int = 30,
    top_n: int = 8,
) -> list[tuple[str, int, float, float]]:
    stats: list[tuple[str, int, float, float]] = []
    for key, delays in bucket.items():
        if len(delays) < min_n:
            continue
        over10 = sum(1 for d in delays if d > 10)
        stats.append((key, len(delays), pct(over10, len(delays)), sum(delays) / len(delays)))
    stats.sort(key=lambda x: (-x[2], -x[3]))
    return stats[:top_n]


def analyze_rows(headers: list[str], rows: list[dict]) -> dict:
    picked_col = find_column(headers, PICKED_ALIASES)
    delivery_col = find_column(headers, DELIVERY_START_ALIASES)
    store_col = find_column(headers, STORE_ALIASES)
    courier_col = find_column(headers, COURIER_ALIASES)
    order_col = find_column(headers, ORDER_ALIASES)

    if not picked_col or not delivery_col:
        raise ValueError(
            f"Не найдены колонки Picked_date / дата начала доставки. "
            f"Заголовки: {headers[:20]}"
        )

    delays: list[float] = []
    by_store: dict[str, list[float]] = defaultdict(list)
    by_courier: dict[str, list[float]] = defaultdict(list)
    by_hour: dict[int, list[float]] = defaultdict(list)
    missing_picked = missing_delivery = negative_or_bad = 0
    instant = 0  # < 1 min

    for row in rows:
        picked = parse_dt(row.get(picked_col))
        delivery = parse_dt(row.get(delivery_col))
        if picked is None:
            missing_picked += 1
            continue
        if delivery is None:
            missing_delivery += 1
            continue
        dm = delay_minutes(picked, delivery)
        if dm is None:
            negative_or_bad += 1
            continue
        delays.append(dm)
        if dm < 1:
            instant += 1
        by_hour[picked.hour].append(dm)
        if store_col:
            store = str(row.get(store_col, "")).strip() or "(без магазина)"
            by_store[store].append(dm)
        if courier_col:
            courier = str(row.get(courier_col, "")).strip() or "(без курьера)"
            by_courier[courier].append(dm)

    total = len(delays)
    thresholds = (5, 10, 15, 20, 30)
    over = {t: sum(1 for d in delays if d > t) for t in thresholds}

    delay_sorted = sorted(delays)
    p50 = delay_sorted[int(total * 0.5)] if total else 0
    p90 = delay_sorted[min(int(total * 0.9), total - 1)] if total else 0
    p95 = delay_sorted[min(int(total * 0.95), total - 1)] if total else 0

    return {
        "picked_col": picked_col,
        "delivery_col": delivery_col,
        "store_col": store_col,
        "courier_col": courier_col,
        "order_col": order_col,
        "rows_total": len(rows),
        "rows_analyzed": total,
        "missing_picked": missing_picked,
        "missing_delivery": missing_delivery,
        "negative_or_bad": negative_or_bad,
        "instant_lt_1min": instant,
        "p50_min": p50,
        "p90_min": p90,
        "p95_min": p95,
        "avg_min": sum(delays) / total if total else 0,
        "over_counts": over,
        "top_stores_over10": top_offenders(by_store),
        "top_couriers_over10": top_offenders(by_courier),
        "worst_hours_p90": sorted(
            [
                (h, len(v), sorted(v)[min(int(len(v) * 0.9), len(v) - 1)] if v else 0)
                for h, v in by_hour.items()
                if len(v) >= 50
            ],
            key=lambda x: -x[2],
        )[:6],
    }


def format_report(path: Path, summary: dict) -> str:
    total = summary["rows_analyzed"] or 1
    over = summary["over_counts"]
    lines = [
        "📊 Задержка: собран (Picked) → начало доставки",
        f"Файл: {path.name}",
        f"Строк в файле: {summary['rows_total']:,}".replace(",", " "),
        f"С валидными датами: {summary['rows_analyzed']:,}".replace(",", " "),
        f"Колонки: «{summary['picked_col']}» → «{summary['delivery_col']}»",
        "",
        "⏱ Задержка (мин), Picked → старт доставки:",
        f"• P50={summary['p50_min']:.1f}  P90={summary['p90_min']:.1f}  "
        f"P95={summary['p95_min']:.1f}  avg={summary['avg_min']:.1f}",
        f"• Мгновенно (<1 мин): {summary['instant_lt_1min']:,} "
        f"({pct(summary['instant_lt_1min'], total):.1f}%)".replace(",", " "),
        "",
        "📈 Доля заказов с задержкой после сборки:",
    ]
    for t in (5, 10, 15, 20, 30):
        cnt = over.get(t, 0)
        lines.append(f"• >{t} мин: {cnt:,} ({pct(cnt, total):.1f}%)".replace(",", " "))

    skip = summary["missing_picked"] + summary["missing_delivery"] + summary["negative_or_bad"]
    if skip:
        lines.append("")
        lines.append(
            f"⚠️ Пропущено: нет Picked={summary['missing_picked']}, "
            f"нет доставки={summary['missing_delivery']}, "
            f"отриц./битые={summary['negative_or_bad']}"
        )

    if summary["top_stores_over10"]:
        lines.append("")
        lines.append("🏪 Магазины — чаще задержка >10 мин (мин. 30 заказов):")
        for name, n, rate, avg in summary["top_stores_over10"]:
            short = name[:48] + ("…" if len(name) > 48 else "")
            lines.append(f"• {short}: {rate:.1f}% >10м, avg={avg:.1f}м, n={n}")

    if summary["top_couriers_over10"]:
        lines.append("")
        lines.append("🛵 Курьеры — чаще задержка >10 мин (мин. 30 заказов):")
        for name, n, rate, avg in summary["top_couriers_over10"]:
            short = str(name)[:40]
            lines.append(f"• {short}: {rate:.1f}% >10м, avg={avg:.1f}м, n={n}")

    if summary["worst_hours_p90"]:
        lines.append("")
        lines.append("🕐 Часы пиковой задержки (P90, ≥50 заказов):")
        for h, n, p90 in summary["worst_hours_p90"]:
            lines.append(f"• {h:02d}:00 — P90={p90:.1f}м, n={n}")

    lines.extend(
        [
            "",
            "💡 Идеи:",
            "1) SLA-алерт: если >10 мин после Picked — пуш курьеру/диспетчеру.",
            "2) Разобрать топ-магазины: очередь на выдаче vs GPS «старт доставки».",
            "3) Курьеры с высоким %>10м — переобучение / проверка статуса в приложении.",
            "4) В часы пикового P90 — усилить диспетчеризацию на точке.",
        ]
    )
    return "\n".join(lines)


def latest_inbox_file() -> Path | None:
    candidates: list[Path] = []
    for pattern in ("*.csv", "*.xlsx"):
        candidates.extend(INBOX.glob(pattern))
    if not candidates:
        return None
    return max(candidates, key=lambda p: p.stat().st_mtime)


def main() -> int:
    env = load_env()
    for key, value in env.items():
        os.environ.setdefault(key, value)

    path_arg = Path(sys.argv[1]) if len(sys.argv) > 1 else None
    target = path_arg if path_arg and path_arg.exists() else latest_inbox_file()
    if not target or not target.exists():
        msg = "В /root/research/inbox нет .csv/.xlsx. Отправьте файл боту или /research."
        tg_send(msg, env=env)
        print(msg)
        return 1

    print(f"analyzing {target}", flush=True)
    try:
        if target.suffix.lower() == ".csv":
            headers, rows = read_csv_rows(target)
        else:
            headers, rows = read_xlsx_rows(target)
        summary = analyze_rows(headers, rows)
    except Exception as exc:
        tg_send(f"Ошибка анализа {target.name}: {exc}", env=env)
        raise

    OUT.mkdir(parents=True, exist_ok=True)
    report_path = OUT / f"picked_delivery_{int(time.time())}.json"
    report_path.write_text(
        json.dumps({**summary, "file": str(target)}, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    msg = format_report(target, summary) + f"\n\nJSON: {report_path}"
    tg_send(msg, env=env)
    print(msg)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
