#!/usr/bin/env python3
"""Analyze delivery click research xlsx (large files) and post summary to Telegram owner."""

from __future__ import annotations

import json
import os
import re
import sys
import time
import urllib.parse
import urllib.request
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("pip install openpyxl", file=sys.stderr)
    raise SystemExit(1)

INBOX = Path("/root/research/inbox")
OUT = Path("/root/data/mlbb/research_reports")
ENV_FILE = Path("/root/.video_bot.env")


def load_env() -> dict[str, str]:
    env: dict[str, str] = {}
    for key, value in os.environ.items():
        if key.startswith("TG_"):
            env[key] = value
    if not ENV_FILE.exists():
        return env
    for line in ENV_FILE.read_text().splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, _, v = line.split("=", 1)
        env[k.strip()] = v.strip()
    return env


def tg_send(text: str, env: dict[str, str] | None = None) -> None:
    env = env or load_env()
    token = env.get("TG_BOT_TOKEN", "")
    chat = env.get("TG_CHAT_ID", "")
    if not token or not chat:
        print("TG_BOT_TOKEN/TG_CHAT_ID missing", file=sys.stderr)
        return
    data = urllib.parse.urlencode({"chat_id": chat, "text": text[:3900]}).encode()
    req = urllib.request.Request(
        f"https://api.telegram.org/bot{token}/sendMessage",
        data=data,
        method="POST",
    )
    with urllib.request.urlopen(req, timeout=30) as resp:
        body = json.loads(resp.read().decode())
    if not body.get("ok"):
        print("TG send failed:", body, file=sys.stderr)


def norm_col(name: str) -> str:
    return re.sub(r"[\s\n\r]+", " ", "_", str(name)).strip().lower()


def find_header_row(ws, max_scan: int = 30) -> tuple[int, dict[str, int]]:
    best_row = 1
    best_score = -1
    best_map: dict[str, int] = {}
    for r in range(1, max_scan + 1):
        row: dict[str, int] = {}
        for c in range(1, min(ws.max_column or 1, 80) + 1):
            v = ws.cell(r, c).value
            if v is None:
                continue
            key = norm_col(str(v))
            if not key:
                continue
            row[key] = c
        score = sum(1 for k in row if k and k not in {"order_id", "courier_id"})
        if score > best_score:
            best_score = score
            best_row = r
            best_map = row
    return best_row, best_map


def cell_to_seconds(v) -> float | None:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        # Excel time fraction or minutes heuristic: small floats are often days
        if 0 < float(v) < 1:
            return float(v) * 86400
        return float(v) * 60 if float(v) < 1000 else float(v)
    if isinstance(v, str):
        s = v.strip().replace(",", ".")
        if not s:
            return None
        if s.count(":") >= 2:
            parts = [int(x) for x in s.split(":")[:3]]
            while len(parts) < 3:
                parts.append(0)
            h, m, sec = parts[0], parts[1], parts[2]
            return h * 3600 + m * 60 + sec
        if s.replace(".", "", 1).isdigit():
            return float(s) * 60
    return None


def duration_from_columns(values: dict[int, object], col_indices: list[int]) -> float | None:
    """Sum interpretable durations in stage columns (seconds), return minutes."""
    total_sec = 0.0
    found = False
    for col in col_indices:
        sec = cell_to_seconds(values.get(col))
        if sec is None:
            continue
        total_sec += sec
        found = True
    return (total_sec / 60.0) if found else None


def analyze_file(path: Path) -> dict:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    header_row, header = find_header_row(ws)
    cols = {col_idx: name for name, col_idx in header.items()}

    # Heuristic stage mapping by Russian keywords in header names
    stage_keywords = {
        "собран": ["собран", "готов", "ready", "assembled"],
        "заявка": ["заявк", "создан", "created"],
        "поиск": ["поиск", "search", "назнач", "assign"],
        "ожид": ["ожид", "wait", "долго"],
        "в_пути": ["в пути", "в пути ", "едет", "в магазин", "забрал"],
        "занят": ["занят", "busy", "сборщик"],
    }
    stage_cols: dict[str, list[int]] = {k: [] for k in stage_keywords}
    for key, col in cols.items():
        for stage, kws in stage_keywords.items():
            if any(kw in key for kw in kws):
                stage_cols[stage].append(col)
                break

    # fallback: columns after G are sequential status timestamps
    status_cols = sorted([c for c in cols if c > 7], key=lambda x: cols[x])[:40]

    durations = {s: [] for s in stage_keywords}
    for stage in durations:
        durations[stage] = []

    order_col = None
    for key in ("order_id", "заказ", "order", "id_заказа"):
        if key in header:
            order_col = header[key]
            break
    if not order_col and header:
        order_col = next(iter(header.values()))

    max_rows = int(os.environ.get("RESEARCH_MAX_ROWS", "0"))
    last_row = ws.max_row or header_row
    if max_rows > 0:
        last_row = min(last_row, header_row + max_rows)

    needed: set[int] = set(status_cols)
    if order_col:
        needed.add(order_col)
    for indices in stage_cols.values():
        needed.update(indices)
    needed_sorted = sorted(needed)

    rows_analyzed = 0
    for r in range(header_row + 1, last_row + 1):
        oid = ws.cell(r, order_col).value if order_col else None
        if oid is None or str(oid).strip() == "":
            continue
        values = {c: ws.cell(r, c).value for c in needed_sorted}

        for stage, indices in stage_cols.items():
            if not indices:
                continue
            m = duration_from_columns(values, indices)
            if m is not None:
                durations[stage].append(m)

        rows_analyzed += 1
        if rows_analyzed % 50000 == 0:
            print(f"  … {rows_analyzed} rows", file=sys.stderr)

    def minutes_list(vals: list[float]) -> dict[str, float]:
        if not vals:
            return {}
        vals_sorted = sorted(vals)
        n = len(vals_sorted)
        return {
            "count": n,
            "p50_min": vals_sorted[int(n * 0.5)] if n else 0,
            "p90_min": vals_sorted[int(n * 0.9)] if n else 0,
            "p95_min": vals_sorted[int(n * 0.95)] if n else 0,
            "avg_min": sum(vals_sorted) / n if n else 0,
        }

    # Problem buckets
    problems = {
        "long_search_courier": durations.get("поиск", []) + durations.get("search", []),
        "long_to_store": durations.get("в пути", []) + durations.get("в_пути ", []),
        "courier_wait_picker": durations.get("ожид", []) + durations.get("wait", []),
        "picker_busy": durations.get("занят", []) + durations.get("busy", []),
    }

    return {
        "file": str(path),
        "sheet": ws.title,
        "rows_analyzed": rows_analyzed,
        "header_row": header_row,
        "stage_stats_minutes": {k: minutes_list(v) for k, v in durations.items()},
        "problems": {k: minutes_list(v) for k, v in problems.items() if v},
        "status_cols_found": status_cols[:12],
    }


def main() -> int:
    env = load_env()
    for key, value in env.items():
        os.environ.setdefault(key, value)

    files = sorted(INBOX.glob("*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
    if not files:
        print("no xlsx in", INBOX)
        return 0

    OUT.mkdir(parents=True, exist_ok=True)
    all_summaries = []

    for path in files:
        try:
            summary = analyze_file(path)
            all_summaries.append(summary)
            print("analyzed", path.name, summary["rows_analyzed"])
        except Exception as exc:
            print("FAIL", path.name, exc, file=sys.stderr)

    if not all_summaries:
        tg_send(
            "В /root/research/inbox/ нет .xlsx. Пришлите боту ссылку transfer.sh или /research <url>",
            env=env,
        )
        return 1

    # aggregate across files (same stage keys)
    from collections import defaultdict

    agg_probs: dict[str, list[float]] = defaultdict(list)
    agg_stats: dict[str, dict] = {}
    total_rows = 0
    for s in all_summaries:
        total_rows += s["rows_analyzed"]
        for k, v in s["problems"].items():
            agg_probs[k].extend(v)
        for stage, st in s["stage_stats_minutes"].items():
            if stage not in agg_stats:
                agg_stats[stage] = {"count": 0}
            agg_stats[stage]["count"] += 1
            agg_stats[stage]["p50"] = agg_stats[stage].get("p50", 0) + st["p50_min"]
            agg_stats[stage]["p90"] = agg_stats[stage].get("p90", 0) + st["p90_min"]
            agg_stats[stage]["avg_min"] = agg_stats[stage].get("avg_min", 0) + st["avg_min"]

    lines = [
        "📊 Анализ доставки (click research)",
        f"Файлов: {len(files)} | Строк: {total_rows}",
        "",
        "Топ проблем (минуты, P50):",
    ]
    for k, st in sorted(agg_probs.items(), key=lambda x: -sum(x), reverse=True)[:8]:
        mins = minutes_list(st)
        lines.append(f"• {k}: P50={mins['p50_min']:.0f}м, P90={mins['p90_min']:.0f}м, n={mins['count']}")

    for stage, st in agg_stats.items():
        lines.append(f"• {stage}: avg {st['avg_min']:.1f}м (n={st['count']})")

    report_path = OUT / f"delivery_report_{int(time.time())}.json"
    report_path.write_text(json.dumps({"summaries": all_summaries, "aggregate": agg_stats}, ensure_ascii=False, indent=2), encoding="utf-8")

    msg = "\n".join(lines) + f"\n\nПолный отчёт: {report_path}"
    tg_send(msg, env=env)
    print("done", report_path)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())